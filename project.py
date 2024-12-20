# This program is written to identify interrogatives in a given transcription.
# The transcripts in this program are in .cha format.

import re
import os
from openpyxl import Workbook, load_workbook


# This function is for counting the total number of parent utterances
def countutterances():

    directory = "C:/Users/Sahana/Desktop/Edinburgh/"
    extension = ".cha"

    # Enumerate files with the specific extension
    files = [f for f in os.listdir(directory) if f.endswith(extension)]

    #reading all transcript files one by one....
    count = 0
    for index, filename in enumerate(files, start=1):

        # Open the transcription file
        filename = directory+filename
        print(filename)
        with open(filename, 'r', encoding = 'utf8') as file:

            # Reading chat transcript
            content = file.read()

            # Setting up a regex pattern to extract mother and father utterances
            pattern = r"^(\*MOT.*|\*FAT.*)$"

            # Find all matching lines
            questions = re.findall(pattern, content, re.MULTILINE)

            # Counting all utterances
            count = count+len(questions)

        print("total utterances = %d" % count)


# This function is for extracting interrogatives and transferring into excel
def main():
    # Specify the directory and file extension
    directory = "C:/Users/Sahana/Desktop/Edinburgh/"
    extension = ".cha"


    # Path to the Excel file
    file_path = "compliteracy.xlsx"


    # Create an Excel sheet if file does not exist
    if not os.path.exists(file_path):

        # creating a new Excel workbook
        wb = Workbook()
        sheet = wb.active

    else:

        # Loading the existing Excel workbook
        wb = load_workbook(file_path)
        sheet = wb.active


    # Enumerate files with the specific extension
    files = [f for f in os.listdir(directory) if f.endswith(extension)]

    for index, filename in enumerate(files, start=1):

        # Open the transcription file
        filename = directory+filename
        print(filename)
        with open(filename, 'r', encoding = 'utf8') as file:
            content = file.read()

            # regex pattern to identify interrogative patterns
            pattern = r"^(\*MOT.*|\*FAT.*)\?.*$"


            # Finding all lines matching the pattern
            questions = re.findall(pattern, content, re.MULTILINE)

            # Print the matches
            # print(questions)


            # Iterate over all questions
            for question in questions:
                # replacing MOT: and FAT: with empty string
                question = re.sub(r"\*MOT:|\*FAT:", "",question)
                print(question)

                # Identifying the next row to update the excel sheet
                next_row = sheet.max_row + 1

                # Adding the question to the excel sheet
                sheet.cell(row=next_row, column=1, value=question)


                # Saving the Excel workbook
                wb.save(file_path)



### Program begins here
if __name__ == "__main__":
    # this function identifies all interrogatives in the transcripts using python regex
    # main()

    # this function computes the total utterances
    countutterances()
